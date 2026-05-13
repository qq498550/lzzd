"""
廉政意见智答系统 - API 路由模块
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Request, Form, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
from urllib.parse import quote
import json
import io
import csv
import base64

from app.models.database import get_db


def get_column_letter(col_idx):
    """将列索引转换为Excel列字母 (1=A, 2=B, ..., 26=Z, 27=AA, ... )"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def status_to_chinese(status):
    """将状态值转换为中文显示"""
    status_map = {
        'processing': '处理中',
        'completed': '已完成',
        'influence_period_ended': '影响期截止',
        'closed': '已关闭'
    }
    return status_map.get(status, status)


def chinese_to_status(text):
    """将中文状态转换为存储值"""
    status_map = {
        '处理中': 'processing',
        '已完成': 'completed',
        '影响期截止': 'influence_period_ended',
        '已关闭': 'closed'
    }
    return status_map.get(text, text if text in ['processing', 'completed', 'influence_period_ended'] else 'completed')
from app.schemas import (
    DisciplineRecordCreate, DisciplineRecordResponse, DisciplineRecordUpdate,
    ViolationRecordCreate, ViolationRecordResponse, ViolationRecordUpdate,
    PetitionReportCreate, PetitionReportResponse, PetitionReportUpdate,
    AnswerTemplateCreate, AnswerTemplateResponse, AnswerTemplateUpdate,
    QueryRequest, QueryResponse, OperationLogCreate, OperationLogResponse
)
from app.services.integrity_service import IntegrityService
from app.core.config import settings

router = APIRouter()

# 用户数据库（内存存储）
users_db = {
    "admin": {
        "username": "admin",
        "password": settings.admin_password,
        "name": "系统管理员",
        "role": "admin"
    }
}


def verify_admin_credentials(username: str, password: str) -> bool:
    """验证管理员账号"""
    if username in users_db and users_db[username]["role"] == "admin":
        return users_db[username]["password"] == password
    return username == settings.admin_username and password == settings.admin_password


def verify_user_credentials(username: str, password: str) -> dict:
    """验证用户账号"""
    if username in users_db and users_db[username]["password"] == password:
        return {"username": username, "role": users_db[username]["role"], "name": users_db[username]["name"]}
    if username == settings.admin_username and password == settings.admin_password:
        return {"username": username, "role": "admin", "name": "系统管理员"}
    return None


def log_operation(db: Session, module: str, action: str, record_id: int, record_name: str, description: str, operator: str = "系统管理员"):
    """记录操作日志"""
    from app.models.database import OperationLog
    try:
        log = OperationLog(
            module=module,
            action=action,
            record_id=record_id,
            record_name=record_name,
            description=description,
            operator=operator
        )
        db.add(log)
        db.commit()
    except Exception as e:
        print(f"[操作日志] 记录失败: {e}")
        db.rollback()


# ==================== 违纪记录管理 ====================
@router.post("/discipline/", response_model=DisciplineRecordResponse, tags=["违纪记录"])
def create_discipline_record(record: DisciplineRecordCreate, db: Session = Depends(get_db)):
    """创建违纪记录"""
    from app.models.database import DisciplineRecord
    db_record = DisciplineRecord(**record.model_dump())
    db.add(db_record)
    db.commit()
    db.refresh(db_record)
    # 记录操作日志
    log_operation(db, "违纪记录", "创建", db_record.id, db_record.name, f"新增违纪记录：{db_record.name}")
    return db_record


@router.get("/discipline/", response_model=List[DisciplineRecordResponse], tags=["违纪记录"])
def list_discipline_records(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """获取违纪记录列表"""
    from app.models.database import DisciplineRecord
    records = db.query(DisciplineRecord).offset(skip).limit(limit).all()
    return records


@router.get("/discipline/{record_id}", response_model=DisciplineRecordResponse, tags=["违纪记录"])
def get_discipline_record(record_id: int, db: Session = Depends(get_db)):
    """获取单条违纪记录"""
    from app.models.database import DisciplineRecord
    record = db.query(DisciplineRecord).filter(DisciplineRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    return record


@router.put("/discipline/{record_id}", response_model=DisciplineRecordResponse, tags=["违纪记录"])
def update_discipline_record(record_id: int, record: DisciplineRecordUpdate, db: Session = Depends(get_db)):
    """更新违纪记录"""
    from app.models.database import DisciplineRecord
    db_record = db.query(DisciplineRecord).filter(DisciplineRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="记录不存在")

    old_name = db_record.name
    update_data = record.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_record, key, value)

    db.commit()
    db.refresh(db_record)
    # 记录操作日志
    new_name = db_record.name or old_name
    log_operation(db, "违纪记录", "修改", db_record.id, new_name, f"修改违纪记录：{new_name}")
    return db_record


@router.delete("/discipline/{record_id}", tags=["违纪记录"])
def delete_discipline_record(record_id: int, db: Session = Depends(get_db)):
    """删除违纪记录"""
    from app.models.database import DisciplineRecord
    db_record = db.query(DisciplineRecord).filter(DisciplineRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="记录不存在")
    record_name = db_record.name
    db.delete(db_record)
    db.commit()
    # 记录操作日志
    log_operation(db, "违纪记录", "删除", record_id, record_name, f"删除违纪记录：{record_name}")
    return {"message": "删除成功"}


# ==================== 违规记录管理 ====================
@router.post("/violation/", response_model=ViolationRecordResponse, tags=["违规记录"])
def create_violation_record(record: ViolationRecordCreate, db: Session = Depends(get_db)):
    """创建违规记录"""
    from app.models.database import ViolationRecord
    db_record = ViolationRecord(**record.model_dump())
    db.add(db_record)
    db.commit()
    db.refresh(db_record)
    # 记录操作日志
    log_operation(db, "违规记录", "创建", db_record.id, db_record.name, f"新增违规记录：{db_record.name}")
    return db_record


@router.get("/violation/", response_model=List[ViolationRecordResponse], tags=["违规记录"])
def list_violation_records(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """获取违规记录列表"""
    from app.models.database import ViolationRecord
    records = db.query(ViolationRecord).offset(skip).limit(limit).all()
    return records


@router.get("/violation/{record_id}", response_model=ViolationRecordResponse, tags=["违规记录"])
def get_violation_record(record_id: int, db: Session = Depends(get_db)):
    """获取单条违规记录"""
    from app.models.database import ViolationRecord
    record = db.query(ViolationRecord).filter(ViolationRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    return record


@router.put("/violation/{record_id}", response_model=ViolationRecordResponse, tags=["违规记录"])
def update_violation_record(record_id: int, record: ViolationRecordUpdate, db: Session = Depends(get_db)):
    """更新违规记录"""
    from app.models.database import ViolationRecord
    db_record = db.query(ViolationRecord).filter(ViolationRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="记录不存在")

    old_name = db_record.name
    update_data = record.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_record, key, value)

    db.commit()
    db.refresh(db_record)
    # 记录操作日志
    new_name = db_record.name or old_name
    log_operation(db, "违规记录", "修改", db_record.id, new_name, f"修改违规记录：{new_name}")
    return db_record


@router.delete("/violation/{record_id}", tags=["违规记录"])
def delete_violation_record(record_id: int, db: Session = Depends(get_db)):
    """删除违规记录"""
    from app.models.database import ViolationRecord
    db_record = db.query(ViolationRecord).filter(ViolationRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="记录不存在")
    record_name = db_record.name
    db.delete(db_record)
    db.commit()
    # 记录操作日志
    log_operation(db, "违规记录", "删除", record_id, record_name, f"删除违规记录：{record_name}")
    return {"message": "删除成功"}


# ==================== 信访举报管理 ====================
@router.post("/petition/", response_model=PetitionReportResponse, tags=["信访举报"])
def create_petition_report(report: PetitionReportCreate, db: Session = Depends(get_db)):
    """创建信访举报记录"""
    from app.models.database import PetitionReport
    db_record = PetitionReport(**report.model_dump())
    db.add(db_record)
    db.commit()
    db.refresh(db_record)
    # 记录操作日志
    log_operation(db, "信访举报", "创建", db_record.id, db_record.name, f"新增信访举报：{db_record.name}")
    return db_record


@router.get("/petition/", response_model=List[PetitionReportResponse], tags=["信访举报"])
def list_petition_reports(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """获取信访举报列表"""
    from app.models.database import PetitionReport
    records = db.query(PetitionReport).offset(skip).limit(limit).all()
    return records


@router.get("/petition/{report_id}", response_model=PetitionReportResponse, tags=["信访举报"])
def get_petition_report(report_id: int, db: Session = Depends(get_db)):
    """获取单条信访举报"""
    from app.models.database import PetitionReport
    report = db.query(PetitionReport).filter(PetitionReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="记录不存在")
    return report


@router.put("/petition/{report_id}", response_model=PetitionReportResponse, tags=["信访举报"])
def update_petition_report(report_id: int, report: PetitionReportUpdate, db: Session = Depends(get_db)):
    """更新信访举报"""
    from app.models.database import PetitionReport
    db_record = db.query(PetitionReport).filter(PetitionReport.id == report_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="记录不存在")

    old_name = db_record.name
    update_data = report.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_record, key, value)

    db.commit()
    db.refresh(db_record)
    # 记录操作日志
    new_name = db_record.name or old_name
    log_operation(db, "信访举报", "修改", db_record.id, new_name, f"修改信访举报：{new_name}")
    return db_record


@router.delete("/petition/{report_id}", tags=["信访举报"])
def delete_petition_report(report_id: int, db: Session = Depends(get_db)):
    """删除信访举报"""
    from app.models.database import PetitionReport
    db_record = db.query(PetitionReport).filter(PetitionReport.id == report_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="记录不存在")
    record_name = db_record.name
    db.delete(db_record)
    db.commit()
    # 记录操作日志
    log_operation(db, "信访举报", "删除", report_id, record_name, f"删除信访举报：{record_name}")
    return {"message": "删除成功"}


# ==================== 答复模板管理 ====================
@router.post("/template/", response_model=AnswerTemplateResponse, tags=["答复模板"])
def create_template(template: AnswerTemplateCreate, db: Session = Depends(get_db)):
    """创建答复模板"""
    from app.models.database import AnswerTemplate
    db_template = AnswerTemplate(**template.model_dump())
    db.add(db_template)
    db.commit()
    db.refresh(db_template)
    # 记录操作日志
    log_operation(db, "答复模板", "创建", db_template.id, db_template.template_code, f"新增答复模板：{db_template.template_name}")
    return db_template


@router.get("/template/", response_model=List[AnswerTemplateResponse], tags=["答复模板"])
def list_templates(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """获取答复模板列表"""
    from app.models.database import AnswerTemplate
    templates = db.query(AnswerTemplate).offset(skip).limit(limit).all()
    return templates


@router.get("/template/{template_id}", response_model=AnswerTemplateResponse, tags=["答复模板"])
def get_template(template_id: int, db: Session = Depends(get_db)):
    """获取单个模板"""
    from app.models.database import AnswerTemplate
    template = db.query(AnswerTemplate).filter(AnswerTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    return template


@router.put("/template/{template_id}", response_model=AnswerTemplateResponse, tags=["答复模板"])
def update_template(template_id: int, template: AnswerTemplateUpdate, db: Session = Depends(get_db)):
    """更新模板"""
    from app.models.database import AnswerTemplate
    db_template = db.query(AnswerTemplate).filter(AnswerTemplate.id == template_id).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="模板不存在")

    old_name = db_template.template_name
    update_data = template.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_template, key, value)

    db.commit()
    db.refresh(db_template)
    # 记录操作日志
    new_name = db_template.template_name or old_name
    log_operation(db, "答复模板", "修改", db_template.id, db_template.template_code, f"修改答复模板：{new_name}")
    return db_template


@router.delete("/template/{template_id}", tags=["答复模板"])
def delete_template(template_id: int, db: Session = Depends(get_db)):
    """删除模板"""
    from app.models.database import AnswerTemplate
    db_template = db.query(AnswerTemplate).filter(AnswerTemplate.id == template_id).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="模板不存在")
    template_name = db_template.template_name
    template_code = db_template.template_code
    db.delete(db_template)
    db.commit()
    # 记录操作日志
    log_operation(db, "答复模板", "删除", template_id, template_code, f"删除答复模板：{template_name}")
    return {"message": "删除成功"}


# ==================== 智能查询核心接口 ====================
@router.post("/query/", response_model=QueryResponse, tags=["智能查询"])
def intelligent_query(query: QueryRequest, db: Session = Depends(get_db)):
    """廉政意见智能查询"""
    service = IntegrityService(db)

    # 检索记录 - 当选择"其他"类型时，需要查询所有类型的记录
    search_results = service.search_person_records(query.name, query.matter_type)

    # 检查姓名是否在数据库中存在（违纪、违规、信访中任一有记录）
    person_exists = len(search_results) > 0

    # 匹配模板
    match_result = service.match_template(search_results, query.matter_type)

    # 生成答复
    if not person_exists:
        # 姓名在数据库中不存在，返回提示信息
        generated_answer = "查询结果为空，请确认输入是否正确。"
        template_code = ""
        template_name = ""
        conclusion = "未找到记录"
    else:
        generated_answer = service.generate_answer(
            match_result["template_content"],
            query.name,
            query.matter_type,
            search_results
        )
        template_code = match_result["template_code"]
        template_name = match_result["template_name"]
        conclusion = match_result["conclusion"]

    # 记录日志
    service.log_query(
        query_name=query.name,
        matter_type=query.matter_type,
        template_code=template_code,
        conclusion=conclusion
    )

    return QueryResponse(
        query_name=query.name,
        matter_type=query.matter_type,
        search_results=search_results,
        matched_template=template_name,
        template_code=template_code,
        conclusion=conclusion,
        generated_answer=generated_answer
    )


@router.get("/query/logs", response_model=List[dict], tags=["智能查询"])
def get_query_logs(limit: int = 100, db: Session = Depends(get_db)):
    """获取所有查询日志"""
    from app.models.database import QueryLog
    logs = db.query(QueryLog).order_by(
        QueryLog.query_time.desc()
    ).limit(limit).all()
    return [
        {
            "id": log.id,
            "query_name": log.query_name,
            "matter_type": log.matter_type,
            "result_template": log.result_template,
            "conclusion": log.conclusion,
            "query_time": log.query_time.isoformat() if log.query_time else None,
            "operator": log.operator
        }
        for log in logs
    ]


@router.get("/query/{name}/history", tags=["智能查询"])
def get_query_history(name: str, limit: int = 10, db: Session = Depends(get_db)):
    """获取某人的查询历史"""
    from app.models.database import QueryLog
    # 使用精确匹配查询姓名
    logs = db.query(QueryLog).filter(
        QueryLog.query_name == name
    ).order_by(QueryLog.query_time.desc()).limit(limit).all()
    return [
        {
            "id": log.id,
            "query_name": log.query_name,
            "matter_type": log.matter_type,
            "result_template": log.result_template,
            "conclusion": log.conclusion,
            "query_time": log.query_time.isoformat() if log.query_time else None,
            "operator": log.operator
        }
        for log in logs
    ]


# ==================== 操作日志管理 ====================
@router.get("/operation/logs", response_model=List[OperationLogResponse], tags=["操作日志"])
def get_operation_logs(limit: int = 100, db: Session = Depends(get_db)):
    """获取所有操作日志"""
    from app.models.database import OperationLog
    logs = db.query(OperationLog).order_by(
        OperationLog.created_at.desc()
    ).limit(limit).all()
    return logs


@router.delete("/operation/logs", tags=["操作日志"])
def clear_operation_logs(db: Session = Depends(get_db)):
    """清空操作日志"""
    from app.models.database import OperationLog
    db.query(OperationLog).delete()
    db.commit()
    return {"message": "操作日志已清空"}


# ==================== 管理员登录验证 ====================
@router.post("/admin/login", tags=["管理员"])
def admin_login(username: str = Form(...), password: str = Form(...)):
    """管理员登录验证"""
    if verify_admin_credentials(username, password):
        return {"success": True, "message": "登录成功"}
    else:
        raise HTTPException(status_code=401, detail="用户名或密码错误")


# ==================== 数据导入导出 ====================
@router.get("/export/discipline", tags=["数据导入导出"])
def export_discipline_records(db: Session = Depends(get_db)):
    """导出违纪记录为Excel"""
    try:
        from app.models.database import DisciplineRecord
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        records = db.query(DisciplineRecord).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "违纪记录"
        
        # 标题行
        ws.cell(row=1, column=1, value='违纪记录导出')
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        
        # 表头
        headers = ['ID', '姓名*', '分公司', '部门', '职务', '处理机构', '问责情况*', '问责时间*', 
                   '有无影响期', '影响期截止日期', '事由*', '状态']
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 数据行
        for r in records:
            ws.append([
                r.id, r.name, r.branch_company or '', r.department or '', r.position or '',
                r.processing_org or '', r.accountability_type or '', 
                r.accountability_date.strftime('%Y-%m-%d') if r.accountability_date else '',
                '有' if r.has_influence_period else '无',
                r.influence_end_date.strftime('%Y-%m-%d') if r.influence_end_date else '',
                r.reason or '', status_to_chinese(r.status)
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).border = thin_border
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('违纪记录.xlsx')}"}
        )
    except Exception as e:
        import traceback
        print(f"[导出违纪] 错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/export/violation", tags=["数据导入导出"])
def export_violation_records(db: Session = Depends(get_db)):
    """导出违规记录为Excel"""
    try:
        from app.models.database import ViolationRecord
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        records = db.query(ViolationRecord).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "违规记录"
        
        # 标题行
        ws.cell(row=1, column=1, value='违规记录导出')
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        
        # 表头
        headers = ['ID', '姓名*', '分公司', '部门', '职务', '处理机构', '问责类型*', '问责时间*', 
                   '有无影响期', '影响期截止日期', '事由*', '状态']
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 数据行
        for r in records:
            ws.append([
                r.id, r.name, r.branch_company or '', r.department or '', r.position or '',
                r.processing_org or '', r.violation_type or '',
                r.violation_date.strftime('%Y-%m-%d') if r.violation_date else '',
                '有' if r.has_influence_period else '无',
                r.influence_end_date.strftime('%Y-%m-%d') if r.influence_end_date else '',
                r.reason or '', status_to_chinese(r.status)
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).border = thin_border
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('违规记录.xlsx')}"}
        )
    except Exception as e:
        import traceback
        print(f"[导出违规] 错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/export/petition", tags=["数据导入导出"])
def export_petition_records(db: Session = Depends(get_db)):
    """导出信访举报记录为Excel"""
    try:
        from app.models.database import PetitionReport
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        records = db.query(PetitionReport).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "信访举报"
        
        # 标题行
        ws.cell(row=1, column=1, value='信访举报记录导出')
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        
        # 表头
        headers = ['ID', '姓名*', '分公司', '部门', '举报日期*', '举报内容*', '核查结果', 
                   '组织是否采信', '有无影响期', '影响期截止日期', '状态']
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 数据行
        for r in records:
            ws.append([
                r.id, r.name, r.branch_company or '', r.department or '',
                r.report_date.strftime('%Y-%m-%d') if r.report_date else '',
                r.report_content or '', r.verification_result or '',
                '是' if r.organization_adoption else ('否' if r.organization_adoption is False else ''),
                '有' if r.has_influence_period else '无',
                r.influence_end_date.strftime('%Y-%m-%d') if r.influence_end_date else '',
                status_to_chinese(r.status)
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).border = thin_border
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('信访举报记录.xlsx')}"}
        )
    except Exception as e:
        import traceback
        print(f"[导出信访] 错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/export/template", tags=["数据导入导出"])
def export_templates(db: Session = Depends(get_db)):
    """导出答复模板为Excel"""
    try:
        from app.models.database import AnswerTemplate
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        records = db.query(AnswerTemplate).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "答复模板"
        
        # 标题行
        ws.cell(row=1, column=1, value='答复模板导出')
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        
        # 表头
        headers = ['ID', '模板编号', '模板名称', '适用场景', '事项类型', '模板内容', '优先级', '是否启用']
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 数据行
        for r in records:
            ws.append([
                r.id, r.template_code, r.template_name, r.scenario_type, r.matter_type,
                r.template_content, r.priority, '是' if r.is_active else '否'
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).border = thin_border
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('答复模板.xlsx')}"}
        )
    except Exception as e:
        import traceback
        print(f"[导出模板] 错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/template/download", tags=["数据导入导出"])
def download_import_template():
    """下载导入模板"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 违纪记录模板
    writer.writerow(['=== 违纪记录导入模板 ==='])
    writer.writerow(['姓名', '部门', '职务', '处理机构', '问责情况', '问责时间 (YYYY-MM-DD)', 
                     '有无影响期 (true/false)', '影响期截止日期 (YYYY-MM-DD)', '事由', '状态 (已完成/处理中)'])
    writer.writerow(['张三', 'XX部门', 'XX职务', 'XX纪委', '党内警告', '2023-01-15', 
                     'true', '2024-01-15', '违反工作纪律', '已完成'])
    writer.writerow([])
    
    # 违规记录模板
    writer.writerow(['=== 违规记录导入模板 ==='])
    writer.writerow(['姓名', '部门', '职务', '处理机构', '问责类型', '问责时间 (YYYY-MM-DD)', 
                     '有无影响期 (true/false)', '影响期截止日期 (YYYY-MM-DD)', '事由', '状态 (已完成/处理中)'])
    writer.writerow(['李四', 'XX部门', 'XX职务', 'XX纪委', '政务记过', '2023-02-20', 
                     'true', '2024-02-20', '违反廉洁纪律', '已完成'])
    writer.writerow([])
    
    # 信访举报模板
    writer.writerow(['=== 信访举报导入模板 ==='])
    writer.writerow(['姓名', '举报日期 (YYYY-MM-DD)', '举报内容', '核查结果', '组织采信 (true/false/null)', '状态 (已完成/处理中/影响期截止)'])
    writer.writerow(['王五', '2023-03-10', '反映收受礼金问题', '未发现相关证据', 'false', '已完成'])
    writer.writerow([])
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('导入模板.csv')}"}
    )


@router.post("/import/discipline", tags=["数据导入导出"])
async def import_discipline_records(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入违纪记录（支持CSV和Excel格式）"""
    from app.models.database import DisciplineRecord

    filename = file.filename.lower()

    if filename.endswith('.xlsx'):
        # Excel格式
        import io
        from openpyxl import load_workbook

        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):  # 跳过标题和表头
            if not row[0]:
                continue
            try:
                has_period = str(row[7]).strip().lower() in ['true', '有'] if row[7] else True
                influence_end = None
                if row[8] and str(row[8]).strip():
                    try:
                        influence_end = datetime.strptime(str(row[8]).strip(), '%Y-%m-%d').date()
                    except:
                        pass

                record = DisciplineRecord(
                    name=str(row[0]).strip(),
                    branch_company=str(row[1]).strip() if row[1] else None,
                    department=str(row[2]).strip() if row[2] else None,
                    position=str(row[3]).strip() if row[3] else None,
                    processing_org=str(row[4]).strip() if row[4] else None,
                    accountability_type=str(row[5]).strip(),
                    accountability_date=datetime.strptime(str(row[6]).strip(), '%Y-%m-%d').date(),
                    has_influence_period=has_period,
                    influence_end_date=influence_end,
                    reason=str(row[9]).strip() if row[9] else '',
                    status=chinese_to_status(str(row[10]).strip()) if row[10] else 'completed'
                )
                db.add(record)
                count += 1
            except Exception as e:
                print(f"[导入违纪] 第{row_idx}行处理失败: {e}")
                continue
    else:
        # CSV格式
        contents = await file.read()
        lines = contents.decode('utf-8').split('\n')

        count = 0
        for i, line in enumerate(lines[1:], start=2):
            if not line.strip():
                continue
            try:
                values = line.split(',')
                if len(values) < 6:
                    continue

                has_period = values[6].strip().lower() in ['true', '有'] if len(values) > 6 else True
                influence_end = None
                if len(values) > 7 and values[7].strip():
                    try:
                        influence_end = datetime.strptime(values[7].strip(), '%Y-%m-%d').date()
                    except:
                        pass

                record = DisciplineRecord(
                    name=values[0].strip(),
                    branch_company=values[1].strip() if len(values) > 1 and values[1].strip() else None,
                    department=values[2].strip() if len(values) > 2 and values[2].strip() else None,
                    position=values[3].strip() if len(values) > 3 and values[3].strip() else None,
                    processing_org=values[4].strip() if len(values) > 4 and values[4].strip() else None,
                    accountability_type=values[5].strip(),
                    accountability_date=datetime.strptime(values[6].strip(), '%Y-%m-%d').date(),
                    has_influence_period=has_period,
                    influence_end_date=influence_end,
                    reason=values[9].strip() if len(values) > 9 and values[9].strip() else '',
                    status=chinese_to_status(values[10].strip()) if len(values) > 10 and values[10].strip() else 'completed'
                )
                db.add(record)
                count += 1
            except Exception as e:
                continue

    db.commit()
    return {"success": True, "message": f"成功导入 {count} 条违纪记录"}


@router.post("/import/violation", tags=["数据导入导出"])
async def import_violation_records(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入违规记录（支持CSV和Excel格式）"""
    from app.models.database import ViolationRecord

    filename = file.filename.lower()

    if filename.endswith('.xlsx'):
        # Excel格式
        import io
        from openpyxl import load_workbook

        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row[0]:
                continue
            try:
                has_period = str(row[7]).strip().lower() in ['true', '有'] if row[7] else True
                influence_end = None
                if row[8] and str(row[8]).strip():
                    try:
                        influence_end = datetime.strptime(str(row[8]).strip(), '%Y-%m-%d').date()
                    except:
                        pass

                record = ViolationRecord(
                    name=str(row[0]).strip(),
                    branch_company=str(row[1]).strip() if row[1] else None,
                    department=str(row[2]).strip() if row[2] else None,
                    position=str(row[3]).strip() if row[3] else None,
                    processing_org=str(row[4]).strip() if row[4] else None,
                    violation_type=str(row[5]).strip(),
                    violation_date=datetime.strptime(str(row[6]).strip(), '%Y-%m-%d').date(),
                    has_influence_period=has_period,
                    influence_end_date=influence_end,
                    reason=str(row[9]).strip() if row[9] else '',
                    status=chinese_to_status(str(row[10]).strip()) if row[10] else 'completed'
                )
                db.add(record)
                count += 1
            except Exception as e:
                print(f"[导入违规] 第{row_idx}行处理失败: {e}")
                continue
    else:
        # CSV格式
        contents = await file.read()
        lines = contents.decode('utf-8').split('\n')

        count = 0
        for i, line in enumerate(lines[1:], start=2):
            if not line.strip():
                continue
            try:
                values = line.split(',')
                if len(values) < 6:
                    continue

                has_period = values[6].strip().lower() in ['true', '有'] if len(values) > 6 else True
                influence_end = None
                if len(values) > 7 and values[7].strip():
                    try:
                        influence_end = datetime.strptime(values[7].strip(), '%Y-%m-%d').date()
                    except:
                        pass

                record = ViolationRecord(
                    name=values[0].strip(),
                    branch_company=values[1].strip() if len(values) > 1 and values[1].strip() else None,
                    department=values[2].strip() if len(values) > 2 and values[2].strip() else None,
                    position=values[3].strip() if len(values) > 3 and values[3].strip() else None,
                    processing_org=values[4].strip() if len(values) > 4 and values[4].strip() else None,
                    violation_type=values[5].strip(),
                    violation_date=datetime.strptime(values[6].strip(), '%Y-%m-%d').date(),
                    has_influence_period=has_period,
                    influence_end_date=influence_end,
                    reason=values[9].strip() if len(values) > 9 and values[9].strip() else '',
                    status=chinese_to_status(values[10].strip()) if len(values) > 10 and values[10].strip() else 'completed'
                )
                db.add(record)
                count += 1
            except Exception as e:
                continue

    db.commit()
    return {"success": True, "message": f"成功导入 {count} 条违规记录"}


@router.post("/import/petition", tags=["数据导入导出"])
async def import_petition_records(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入信访举报记录（支持CSV和Excel格式）"""
    from app.models.database import PetitionReport

    filename = file.filename.lower()

    if filename.endswith('.xlsx'):
        # Excel格式
        import io
        from openpyxl import load_workbook

        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row[0]:
                continue
            try:
                org_adoption = None
                if row[6] and str(row[6]).strip():
                    if str(row[6]).strip().lower() in ['true', '是']:
                        org_adoption = True
                    elif str(row[6]).strip().lower() in ['false', '否']:
                        org_adoption = False

                has_period = str(row[7]).strip().lower() in ['true', '有'] if row[7] else False
                influence_end = None
                if row[8] and str(row[8]).strip():
                    try:
                        influence_end = datetime.strptime(str(row[8]).strip(), '%Y-%m-%d').date()
                    except:
                        pass

                record = PetitionReport(
                    name=str(row[0]).strip(),
                    branch_company=str(row[1]).strip() if row[1] else None,
                    department=str(row[2]).strip() if row[2] else None,
                    report_date=datetime.strptime(str(row[3]).strip(), '%Y-%m-%d').date(),
                    report_content=str(row[4]).strip() if row[4] else '',
                    verification_result=str(row[5]).strip() if row[5] else None,
                    organization_adoption=org_adoption,
                    has_influence_period=has_period,
                    influence_end_date=influence_end,
                    status=chinese_to_status(str(row[9]).strip()) if row[9] else 'processing'
                )
                db.add(record)
                count += 1
            except Exception as e:
                print(f"[导入信访] 第{row_idx}行处理失败: {e}")
                continue
    else:
        # CSV格式
        contents = await file.read()
        lines = contents.decode('utf-8').split('\n')

        count = 0
        for i, line in enumerate(lines[1:], start=2):
            if not line.strip():
                continue
            try:
                values = line.split(',')
                if len(values) < 3:
                    continue

                org_adoption = None
                if len(values) > 4 and values[4].strip():
                    if values[4].strip().lower() in ['true', '是']:
                        org_adoption = True
                    elif values[4].strip().lower() in ['false', '否']:
                        org_adoption = False

                has_period = len(values) > 6 and values[6].strip().lower() in ['true', '有']
                influence_end = None
                if len(values) > 7 and values[7].strip():
                    try:
                        influence_end = datetime.strptime(values[7].strip(), '%Y-%m-%d').date()
                    except:
                        pass

                record = PetitionReport(
                    name=values[0].strip(),
                    branch_company=values[1].strip() if len(values) > 1 and values[1].strip() else None,
                    department=values[2].strip() if len(values) > 2 and values[2].strip() else None,
                    report_date=datetime.strptime(values[3].strip(), '%Y-%m-%d').date(),
                    report_content=values[4].strip() if len(values) > 4 else '',
                    verification_result=values[5].strip() if len(values) > 5 and values[5].strip() else None,
                    organization_adoption=org_adoption,
                    has_influence_period=has_period,
                    influence_end_date=influence_end,
                    status=chinese_to_status(values[9].strip()) if len(values) > 9 and values[9].strip() else 'processing'
                )
                db.add(record)
                count += 1
            except Exception as e:
                continue

    db.commit()
    return {"success": True, "message": f"成功导入 {count} 条信访举报记录"}


# ==================== 用户管理 ====================
@router.get("/admin/users", tags=["用户管理"])
def get_users(db: Session = Depends(get_db)):
    """获取所有用户列表（仅管理员）"""
    user_list = []
    for username, info in users_db.items():
        user_list.append({
            "username": username,
            "name": info.get("name", username),
            "role": info["role"]
        })
    return user_list


@router.post("/admin/users", tags=["用户管理"])
def create_user(username: str = Form(...), password: str = Form(...), 
                name: str = Form(...), role: str = Form("user")):
    """创建新用户（仅管理员）"""
    if username in users_db:
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    users_db[username] = {
        "password": password,
        "name": name,
        "role": role
    }
    return {"success": True, "message": f"用户 {username} 创建成功"}


@router.delete("/admin/users/{username}", tags=["用户管理"])
def delete_user(username: str):
    """删除用户（仅管理员）"""
    if username not in users_db:
        raise HTTPException(status_code=404, detail="用户不存在")
    if username == "admin":
        raise HTTPException(status_code=400, detail="不能删除管理员账号")
    
    del users_db[username]
    return {"success": True, "message": f"用户 {username} 已删除"}


@router.post("/admin/users/{username}/reset-password", tags=["用户管理"])
def reset_password(username: str, new_password: str = Form(...)):
    """重置用户密码（仅管理员）"""
    if username not in users_db:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    users_db[username]["password"] = new_password
    return {"success": True, "message": f"用户 {username} 密码已重置"}


@router.post("/admin/change-password", tags=["管理员"])
def change_password(old_password: str = Form(...), new_password: str = Form(...)):
    """修改管理员密码"""
    # 这里简单处理，实际应该结合 session 或 token 验证当前登录用户
    if not verify_admin_credentials("admin", old_password):
        raise HTTPException(status_code=401, detail="原密码错误")
    
    users_db["admin"]["password"] = new_password
    return {"success": True, "message": "密码修改成功"}


@router.post("/user/login", tags=["用户"])
def user_login(username: str = Form(...), password: str = Form(...)):
    """用户登录验证"""
    user = verify_user_credentials(username, password)
    if user:
        return {"success": True, "message": "登录成功", "user": user}
    else:
        raise HTTPException(status_code=401, detail="用户名或密码错误")


# ==================== PDF 导出 ====================
@router.post("/export/pdf", response_class=HTMLResponse, tags=["数据导入导出"])
def export_pdf_answer(query_name: str = Form(...), matter_type: str = Form(...), 
                      generated_answer: str = Form(...), conclusion: str = Form(...)):
    """导出答复函为PDF（使用浏览器打印功能）"""
    html_content = f"""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <title>廉政意见答复函</title>
        <style>
            body {{ font-family: 'SimSun', '宋体', serif; padding: 40px; line-height: 1.8; }}
            .header {{ text-align: center; margin-bottom: 40px; }}
            .title {{ font-size: 24px; font-weight: bold; margin-bottom: 10px; }}
            .content {{ font-size: 16px; text-indent: 2em; margin: 30px 0; }}
            .conclusion {{ font-size: 16px; margin-top: 30px; }}
            .footer {{ margin-top: 60px; text-align: right; }}
            .date {{ margin-top: 20px; }}
            @media print {{
                body {{ padding: 20px; }}
                .no-print {{ display: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">廉政意见答复函</div>
        </div>
        <div class="content">
            {generated_answer.replace(chr(10), '<br>').replace(' ', '&nbsp;&nbsp;')}
        </div>
        <div class="conclusion">
            <strong>结论：{conclusion}</strong>
        </div>
        <div class="footer">
            <div>陕西销售纪委</div>
            <div class="date">{datetime.now().strftime('%Y年%m月%d日')}</div>
        </div>
        <div class="no-print" style="margin-top: 40px; text-align: center;">
            <button onclick="window.print()" style="padding: 10px 30px; font-size: 16px; cursor: pointer;">🖨️ 打印/保存为PDF</button>
            <button onclick="window.close()" style="padding: 10px 30px; font-size: 16px; cursor: pointer; margin-left: 20px;">关闭</button>
        </div>
        <script>
            // 自动触发打印对话框
            // window.onload = function() {{ window.print(); }}
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)

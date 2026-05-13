"""生成Excel导入模板文件"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os


def create_template(filename, headers, sample_data, instructions):
    """创建带样式和示例的Excel模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据导入"

    # 设置列宽
    for i, header in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 18

    # 标题样式
    title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')

    # 说明区域
    ws.merge_cells('A1:' + get_column_letter(len(headers)) + '1')
    ws['A1'] = instructions
    ws['A1'].font = Font(name='微软雅黑', size=10, italic=True, color='666666')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40

    # 表头
    header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for i, header in enumerate(headers, 2):
        cell = ws.cell(row=2, column=i - 1, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[2].height = 25

    # 示例数据行
    sample_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    sample_font = Font(name='微软雅黑', size=10, color='999999')

    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = sample_font
            cell.fill = sample_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_idx].height = 22

    # 添加空的数据行（用于填写）
    empty_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    empty_font = Font(name='微软雅黑', size=10, color='000000')

    for row_idx in range(3, 13):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = empty_fill
            cell.font = empty_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 边框样式
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    for row in ws.iter_rows(min_row=2, max_row=12, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # 冻结首行
    ws.freeze_panes = 'A3'

    wb.save(filename)
    print(f"已生成模板: {filename}")


def main():
    # 确保templates目录存在
    templates_dir = os.path.join(os.path.dirname(__file__), 'app', 'templates')
    os.makedirs(templates_dir, exist_ok=True)

    # 1. 违纪记录模板
    discipline_headers = ['姓名*', '分公司', '部门', '职务', '处理机构',
                          '问责情况*', '问责时间*', '有无影响期', '影响期截止日期', '事由*', '状态']
    discipline_sample = [
        ['张三', '第一分公司', '财务部', '经理', '纪委', '警告', '2024-01-15', '有', '2025-01-14', '违反财务规定', 'completed'],
        ['李四', '第二分公司', '销售部', '', '组织部', '记过', '2024-03-20', '有', '2026-03-19', '业绩造假', 'completed'],
    ]
    discipline_instructions = "【违纪记录导入模板】*号为必填项。有无影响期填：有/无。状态填：completed/processing。日期格式：YYYY-MM-DD"
    create_template(
        os.path.join(templates_dir, 'discipline_template.xlsx'),
        discipline_headers,
        discipline_sample,
        discipline_instructions
    )

    # 2. 违规记录模板
    violation_headers = ['姓名*', '分公司', '部门', '职务', '处理机构',
                         '问责类型*', '问责时间*', '有无影响期', '影响期截止日期', '事由*', '状态']
    violation_sample = [
        ['王五', '第三分公司', '采购部', '主管', '监察部', '严重警告', '2023-06-10', '有', '2025-06-09', '违规采购', 'completed'],
        ['赵六', '第一分公司', '人事部', '', '组织部', '诫勉谈话', '2024-02-28', '无', '', '考勤违规', 'completed'],
    ]
    violation_instructions = "【违规记录导入模板】*号为必填项。有无影响期填：有/无。状态填：completed/processing。日期格式：YYYY-MM-DD"
    create_template(
        os.path.join(templates_dir, 'violation_template.xlsx'),
        violation_headers,
        violation_sample,
        violation_instructions
    )

    # 3. 信访举报模板
    petition_headers = ['姓名*', '分公司', '部门', '举报日期*', '举报内容*',
                         '核查结果', '组织是否采信', '有无影响期', '影响期截止日期', '状态']
    petition_sample = [
        ['孙七', '第二分公司', '市场部', '2024-05-01', '反映贪污受贿问题', '查无实据', '是', '无', '', 'completed'],
        ['周八', '第一分公司', '技术部', '2024-06-15', '反映违规操作', '', '', '无', '', 'processing'],
    ]
    petition_instructions = "【信访举报导入模板】*号为必填项。组织是否采信填：是/否/空。有无影响期填：有/无/空。状态填：processing/completed/influence_period_ended。日期格式：YYYY-MM-DD"
    create_template(
        os.path.join(templates_dir, 'petition_template.xlsx'),
        petition_headers,
        petition_sample,
        petition_instructions
    )

    print("\n所有模板文件已生成完毕！")
    print("模板位置: app/templates/")


if __name__ == '__main__':
    main()
